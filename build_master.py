import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

BASE = r"C:\Users\Administrator\Documents\department_history"

# ---------- load dim_org_unit.csv ----------
name_by_code = {}
type_by_code = {}
parent_by_code = {}
temp_codes = set()
dept_rows = []            # (code, name) for unit_type == department, in file order
majors_by_parent = {}     # dept_code -> list of (major_code, major_name), in file order

with open(f"{BASE}\\dim_org_unit.csv", encoding="utf-8-sig") as f:
    r = csv.DictReader(f)
    for row in r:
        code = row["unit_code"].strip()
        name = row["unit_name"].strip()
        utype = row["unit_type"].strip()
        parent = row["parent_code"].strip()
        name_by_code[code] = name
        type_by_code[code] = utype
        parent_by_code[code] = parent
        if row.get("is_temp_code", "").strip() == "1":
            temp_codes.add(code)
        if utype == "department":
            dept_rows.append((code, name))

for code, utype in type_by_code.items():
    if utype == "major":
        parent = parent_by_code.get(code, "")
        if parent and type_by_code.get(parent) == "department":
            majors_by_parent.setdefault(parent, []).append((code, name_by_code[code]))

# ---------- load org_unit_relation.csv ----------
edges_from_dept = {}  # dept_code -> list of relation dict
relations = []
note_by_target = {}  # after_dept_code -> note (from the CSV 'note' column, if present)
with open(f"{BASE}\\org_unit_relation.csv", encoding="utf-8-sig") as f:
    r = csv.DictReader(f)
    for row in r:
        rel = {k: v.strip() for k, v in row.items()}
        relations.append(rel)
        pd = rel["prev_dept_code"]
        if pd:
            edges_from_dept.setdefault(pd, []).append(rel)
        note = rel.get("note", "")
        after = rel["after_dept_code"]
        if note and after:
            note_by_target[after] = note


def rows_for(dept_code, major_code):
    """Relation rows continuing from (dept_code, major_code), tracking the
    major-track so unrelated sibling branches don't get merged together."""
    rows = edges_from_dept.get(dept_code, [])
    exact = [r for r in rows if r["prev_major_code"] == major_code]
    if exact:
        return exact
    if major_code != "":
        # fall back to a blanket whole-department transition, if any
        blanket = [r for r in rows if r["prev_major_code"] == ""]
        if blanket:
            return blanket
    return []


def format_pair(dept_code, major_code):
    """학부(전공) 표기: 전공이 있으면 상위 학부명(전공명), 없으면 학부/학과명 그대로."""
    dept_name = name_by_code.get(dept_code, dept_code)
    if major_code:
        major_name = name_by_code.get(major_code, major_code)
        return f"{dept_name}({major_name})"
    return dept_name


def resolve_node(dept_code, major_code, visited, closed_hit):
    """Follow a single (dept, major) track forward; returns set of terminal (dept, major) pairs."""
    key = (dept_code, major_code)
    if key in visited:
        return set()
    visited = visited | {key}
    rows = rows_for(dept_code, major_code)
    if not rows:
        return {(dept_code, major_code)}
    targets = set()
    for rel in rows:
        after = rel["after_dept_code"]
        if not after:
            closed_hit.append((dept_code, major_code, rel["change_year"], rel["valid_until"]))
            continue
        targets |= resolve_node(after, rel["after_major_code"], visited, closed_hit)
    return targets


def resolve_dept_aggregate(dept_code):
    """학부/학과 단위 관점: 그 아래 있었던 모든 전공 트랙을 합쳐서 계승 관계를 본다."""
    closed_hit = []
    rows = edges_from_dept.get(dept_code, [])
    if not rows:
        return {(dept_code, "")}, closed_hit
    majors = sorted(set(r["prev_major_code"] for r in rows))
    targets = set()
    for m in majors:
        targets |= resolve_node(dept_code, m, set(), closed_hit)
    return targets, closed_hit


def build_row(label, own_code, own_major, targets, closed_hit, is_never_changed):
    target_labels = sorted(set(format_pair(d, m) for d, m in targets))
    remark_parts = []
    b_value = ""

    if is_never_changed:
        b_value = label
        if own_code in temp_codes or own_major in temp_codes:
            remark_parts.append("가코드(임시코드) - 정식 학과명 확정 전")
    elif targets:
        if len(target_labels) == 1:
            b_value = target_labels[0]
        else:
            b_value = " / ".join(target_labels)
            remark_parts.append(f"분할됨(여러 단위로 계승): {', '.join(target_labels)} - 매칭 확인 필요")
        if closed_hit:
            remark_parts.append("일부 트랙은 도중에 폐과 처리됨")
        temp_hit = [format_pair(d, m) for d, m in targets if d in temp_codes or m in temp_codes]
        if temp_hit:
            remark_parts.append(f"가코드(임시코드) 경유: {', '.join(sorted(set(temp_hit)))} - 정식 학과명 확정 전")
    else:
        years = sorted(set(y for _, _, y, _ in closed_hit))
        remark_parts.append(f"폐과({'/'.join(years)}년) - 후속 단위 없음")

    note_codes = sorted(set(
        c for d, m in targets for c in (d, m) if c in note_by_target
    )) if not is_never_changed else (
        [c for c in (own_code, own_major) if c in note_by_target]
    )
    note = "\n\n".join(note_by_target[c] for c in note_codes) if note_codes else None

    return (label, b_value, "; ".join(remark_parts), note)


results = []
for dept_code, dept_name in dept_rows:
    # 1) 학부/학과 단위 자체의 이력 (하위 전공을 모두 합쳐서 봄)
    agg_targets, agg_closed = resolve_dept_aggregate(dept_code)
    never_changed = not edges_from_dept.get(dept_code)
    results.append(build_row(dept_name, dept_code, "", agg_targets, agg_closed, never_changed))

    # 2) 그 학부 아래에 있었던 전공 단위 각각의 이력 (학부(전공) 표기, 트랙별 개별 추적)
    for major_code, major_name in majors_by_parent.get(dept_code, []):
        label = format_pair(dept_code, major_code)
        closed_hit = []
        targets = resolve_node(dept_code, major_code, set(), closed_hit)
        never_changed_m = not rows_for(dept_code, major_code)
        results.append(build_row(label, dept_code, major_code, targets, closed_hit, never_changed_m))

# ---------- write xlsx ----------
wb = Workbook()
ws = wb.active
ws.title = "학과 매칭 마스터"

headers = ["A: 과거 학과/전공명", "B: 2026학년도 기준 명칭", "C: 비고"]
ws.append(headers)
for col in range(1, 4):
    c = ws.cell(row=1, column=col)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, (label, b_value, remark, note) in enumerate(results, start=2):
    ws.append([label, b_value, remark])
    if note:
        b_cell = ws.cell(row=row_idx, column=2)
        b_cell.comment = Comment(note, "학과 이력 매칭")

widths = [34, 40, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:C{len(results)+1}"

out_path = f"{BASE}\\학과_매칭_마스터파일.xlsx"
wb.save(out_path)
print("saved:", out_path)
print("total rows:", len(results))
ambiguous = [x for x in results if x[2]]
print("flagged:", len(ambiguous))
noted = [x for x in results if x[3]]
print("tooltip 부착:", len(noted))
